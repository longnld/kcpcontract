from unicodedata import category
from urllib import request
import openpyxl
from datetime import datetime
from excelExtract.models import document,excel,pdfFile,accountEmail,excelAccount
import os
from django.core.files import File
# import json
# import pdfkit 
from user.models import Profile
from fpdf import FPDF, HTMLMixin
from datetime import datetime
import re
class PDF(FPDF, HTMLMixin):
	# FPDF("L", "mm", "A4")
	
	pass
def importDataExcel(path, user=None):
	wb=openpyxl.load_workbook(path,data_only=True)
	file=document(document=path)
	if user != None:
		file.upload_by = user
	file.save()
	#TÌM HEADERS,LOẠI CT TRONG SHEET CẦN TÌM/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
	programCate=[]	#LIST TỔNG HỢP LẠI CHƯƠNG TRÌNH CỦA 2 BIẾN DƯỚI (ĐỀ PHÒNG TRƯỜNG HỢP CÓ LOẠI CT MỚI MÀ SHEET KHÁC KHÔNG CÓ)
	programCate1=[] #CHỨA CÁC LOẠI CT TRONG Ecom Promotion Plan BCC_for SO
	programCate2=[] #CHỨA CÁC LOẠI CT TRONG  MnB Promotion Plan BCC_for SO
	listHeader=[] 	#CHỨA HEADER của Ecom VÀ Mnb Promotion Plan BCC_for SO (HIỆN TẠI DỰA VÀO ECOM ĐỂ XÁC ĐỊNH VÌ CẢ 2 GIỐNG NHAU )
	
	for f in wb.sheetnames:
		
		if f == "Promotion Plan BCC":
			wb.active=wb[f]
			ws=wb.active			
			programCateCol=ws['BQ'] # CỘT LOẠI CT
			group=ws['A'] # CỘT GROUP
			lineEnd=0 # XÁC ĐỊNH ROWS DỮ LIỆU  KẾT THÚC TẠI DÒNG NÀO ,DỰA TRÊN GROUP
			count=0 # ĐẾM SỐ LƯỢNG DATA IMPORT
			#LIỆT KÊ LOẠI CT
			for i in range(3,len(programCateCol)): 
				if  programCateCol[i].value not in programCate1 :
					programCate1.append(programCateCol[i].value)
					if  programCateCol[i].value==None:
						programCate1.pop()
						break
			# XÁC ĐỊNH ROWS DỮ LIỆU  KẾT THÚC TẠI DÒNG NÀO ,DỰA TRÊN GROUP
			for i in range(3,len(group)):		
				if group[i].value != None:
					count=count+1
			lineEnd=count+3
			print("lineEnd" +"{}: {}".format(str(f),lineEnd))
			#import data
			rangeline=lineEnd
			
			for i,row in enumerate(ws.iter_rows(min_row=4,max_col=70,values_only=True),start=4):
				
				if i <= rangeline:
							
						standart=str(row[1]).replace(" ","").lower()
						acc=excelAccount.objects.filter(standardName=standart)
						if acc:
							try:
								excel.objects.create(filename=file,group=row[0],account=acc[0],postStartDate=row[4],postEndDate=row[5],product=row[10],mechanicsGetORDiscount=row[12],noiDungChuongTrinh=row[57],budgetRir=row[59],loaiCt=row[68])	
							except:
								if isinstance(row[4],str) and isinstance(row[5],str):
									excel.objects.create(filename=file,group=row[0],account=acc[0],product=row[10],mechanicsGetORDiscount=row[12],noiDungChuongTrinh=row[57],budgetRir=row[59],loaiCt=row[68])	
								elif isinstance(row[4],str):
									excel.objects.create(filename=file,group=row[0],account=acc[0],postEndDate=row[5],product=row[10],mechanicsGetORDiscount=row[12],noiDungChuongTrinh=row[57],budgetRir=row[59],loaiCt=row[68])	
								elif isinstance(row[5],str):
									excel.objects.create(filename=file,group=row[0],account=acc[0],postStartDate=row[4],product=row[10],mechanicsGetORDiscount=row[12],noiDungChuongTrinh=row[57],budgetRir=row[59],loaiCt=row[68])	
						if not acc :
							print(i,row[1],standart)
							excelAccount.objects.create(account=str(row[1]),standardName=standart)
							try:
								excel.objects.create(filename=file,group=row[0],account=row[1],postStartDate=row[4],postEndDate=row[5],product=row[10],mechanicsGetORDiscount=row[12],noiDungChuongTrinh=row[57],budgetRir=row[59],loaiCt=row[68])
							except:	
								if isinstance(row[4],str) and isinstance(row[5],str):
									excel.objects.create(filename=file,group=row[0],account=row[1],product=row[10],mechanicsGetORDiscount=row[12],noiDungChuongTrinh=row[57],budgetRir=row[59],loaiCt=row[68])	
								elif isinstance(row[4],str):
									excel.objects.create(filename=file,group=row[0],account=row[1],postEndDate=row[5],product=row[10],mechanicsGetORDiscount=row[12],noiDungChuongTrinh=row[57],budgetRir=row[59],loaiCt=row[68])	
								elif isinstance(row[5],str):
									excel.objects.create(filename=file,group=row[0],account=row[1],postStartDate=row[4],product=row[10],mechanicsGetORDiscount=row[12],noiDungChuongTrinh=row[57],budgetRir=row[59],loaiCt=row[68])	
				else:
					break
			
		if f == "Promotion Plan FEM":
			wb.active=wb[f]
			ws=wb.active			
			programCateCol=ws['BQ'] # CỘT LOẠI CT
			group=ws['A'] # CỘT GROUP
			lineEnd=0 # XÁC ĐỊNH ROWS DỮ LIỆU  KẾT THÚC TẠI DÒNG NÀO ,DỰA TRÊN GROUP
			count=0 # ĐẾM SỐ LƯỢNG DATA IMPORT
			#LIỆT KÊ LOẠI CT
			for i in range(3,len(programCateCol)): 
				if  programCateCol[i].value not in programCate1 :
					programCate1.append(programCateCol[i].value)
					if  programCateCol[i].value==None:
						programCate1.pop()
						break
			# XÁC ĐỊNH ROWS DỮ LIỆU  KẾT THÚC TẠI DÒNG NÀO ,DỰA TRÊN GROUP
			for i in range(3,len(group)):		
				if group[i].value != None:
					count=count+1
			lineEnd=count+3
			print("lineEnd" +"{}: {}".format(str(f),lineEnd))
			#import data
			rangeline=lineEnd
			for i,row in enumerate(ws.iter_rows(min_row=4,max_col=70,values_only=True),start=4):
				
				if i <= rangeline:
							
						standart=str(row[1]).replace(" ","").lower()
						acc=excelAccount.objects.filter(standardName=standart)
						print(acc)
						if acc:
							try:
								excel.objects.create(filename=file,group=row[0],account=acc[0],postStartDate=row[4],postEndDate=row[5],product=row[10],mechanicsGetORDiscount=row[12],noiDungChuongTrinh=row[57],budgetRir=row[59],loaiCt=row[68])	
							except:
								if isinstance(row[4],str) and isinstance(row[5],str):
									excel.objects.create(filename=file,group=row[0],account=acc[0],product=row[10],mechanicsGetORDiscount=row[12],noiDungChuongTrinh=row[57],budgetRir=row[59],loaiCt=row[68])	
								elif isinstance(row[4],str):
									excel.objects.create(filename=file,group=row[0],account=acc[0],postEndDate=row[5],product=row[10],mechanicsGetORDiscount=row[12],noiDungChuongTrinh=row[57],budgetRir=row[59],loaiCt=row[68])	
								elif isinstance(row[5],str):
									excel.objects.create(filename=file,group=row[0],account=acc[0],postStartDate=row[4],product=row[10],mechanicsGetORDiscount=row[12],noiDungChuongTrinh=row[57],budgetRir=row[59],loaiCt=row[68])	
						if not acc :
							print(i,row[1],standart)
							excelAccount.objects.create(account=row[1],standardName=standart)
							try:
								excel.objects.create(filename=file,group=row[0],account=row[1],postStartDate=row[4],postEndDate=row[5],product=row[10],mechanicsGetORDiscount=row[12],noiDungChuongTrinh=row[57],budgetRir=row[59],loaiCt=row[68])
							except:	
								if isinstance(row[4],str) and isinstance(row[5],str):
									excel.objects.create(filename=file,group=row[0],account=row[1],product=row[10],mechanicsGetORDiscount=row[12],noiDungChuongTrinh=row[57],budgetRir=row[59],loaiCt=row[68])	
								elif isinstance(row[4],str):
									excel.objects.create(filename=file,group=row[0],account=row[1],postEndDate=row[5],product=row[10],mechanicsGetORDiscount=row[12],noiDungChuongTrinh=row[57],budgetRir=row[59],loaiCt=row[68])	
								elif isinstance(row[5],str):
									excel.objects.create(filename=file,group=row[0],account=row[1],postStartDate=row[4],product=row[10],mechanicsGetORDiscount=row[12],noiDungChuongTrinh=row[57],budgetRir=row[59],loaiCt=row[68])	
				else:
					break
			
			
	# TỔNG HỢP LẠI LOẠI CHƯƠNG TRÌNH CỦA 2 LIST (ĐỀ PHÒNG TRƯỜNG HỢP CÓ LOẠI CT MỚI MÀ LISTS KHÁC KHÔNG CÓ)
	for i in programCate1:
		if i not in programCate:
			programCate.append(i)
	for i in programCate2:
		if i not in programCate:
			programCate.append(i)
			
	print("success")
	return listHeader
	#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
#
# excelExtract.exportFiles(loaict,fileID,loaiAccount)
def exportFiles(loaict,fileID,loaiAccount,user):
	annouceExist=""
	#XÁC ĐỊNH LOẠI CT  FILE
	listloaict=[]
	listAccount=[]
	file=document.objects.get(pk=int(fileID))
	print(file)
	for f in excel.objects.filter(filename=file):
		if f.loaiCt not in listloaict:
			listloaict.append(f.loaiCt)
		if f.account not in listAccount:
			listAccount.append(f.account)	
	print(listloaict)
	print(listAccount)
	# print(listNoiDungCt)
	print(type(loaiAccount),loaiAccount)
	print(type(loaict),loaict)
	print(type(loaiAccount.split(',')),loaiAccount.split(','))

	if loaiAccount=="all" and loaict=="all": 
		
		for f in listAccount:
			emptyList=[]
			errorFlag=False

			pdfAccount=excelAccount.objects.get(account=f)
			try:
				pdfEmail=accountEmail.objects.filter(account=pdfAccount)
			except:
				pdfEmail=[]
			for chuongTrinh in listloaict:
				if excel.objects.filter(filename=file,account=f,loaiCt=chuongTrinh).exists():
					emptyList=[]
					break
				else:
					emptyList.append("{}".format(str(chuongTrinh)))

			if emptyList != []:
				annouceExist=annouceExist+" {}_no_Data ".format(str(f))
				continue
			
			currentAccountCates=excel.objects.filter(filename=file,account=f).values("loaiCt").distinct()
			listloaict2=[]
			categoryString=""
			for i,b in enumerate(currentAccountCates):
				listloaict2.append(b['loaiCt'])
			#filter the categoryString
			for index,categories in enumerate(currentAccountCates):
				removeCategoryFlag=False
				#check if all data is wrong format
				if all(data is None for data in excel.objects.filter(filename=file,account=f,loaiCt=categories['loaiCt']).values_list("postStartDate",flat=True)):
					removeCategoryFlag=True
				if all(data is None for data in excel.objects.filter(filename=file,account=f,loaiCt=categories['loaiCt']).values_list("postEndDate",flat=True)):
					removeCategoryFlag=True
				if removeCategoryFlag:
					listloaict2.remove("{}".format(str(categories['loaiCt'])))
				else:
					if index > 0 :
						categoryString+=",{}".format(categories['loaiCt'])
					else:
						categoryString+="{}".format(categories['loaiCt'])

			if listloaict2 == []:
				annouceExist=annouceExist+" {}_no_Data ".format(str(f)) #empty list 2nd
				continue

			title = 'THÔNG BÁO VỀ CHƯƠNG TRÌNH KHUYẾN MÃI' #.encode('utf-8')
			date_year = "Tp.HCM, Ngày {}".format(str(datetime.now().date().strftime("%d/%m/%Y")))
			pdf = PDF()
			pdf.add_font('arial','',r"static/Fonts/arial.ttf",uni=True)
			pdf.add_font('arial','B',r"static/Fonts/FontsFree-Net-arial-bold.ttf",uni=True)
			pdf.add_font('arial','I',r"static/Fonts/Arial-Italic.ttf",uni=True)
			pdf.set_font('arial','', size=10)
			# pdf=FPDF("L", "mm", "A4")

			pdf.add_page("L")
			title_w = pdf.get_string_width(title) + 6

			doc_w = pdf.w
			center = (doc_w - title_w) / 2
			#tittle
			pdf.text(center,15,title)
			#logo
			pdf.image('static/image/kimberlylogo.jpg', x = 5, w = 60,h=10, y=5,type="jpg")
			# date year

			pdf.set_font('arial','', size=8)
			pdf.text(255,20,date_year)
			# Kính gửi
			pdf.set_font('arial', 'B', size=9)
			text = 'Kính gửi: Quý Khách Hàng Kênh Hiện Đại'

			pdf.text(10, 25, text)
			# Tên chương trình
			pdf.set_font('arial','', size=8)
			text = "Công ty TNHH Kimberly-Clark Việt Nam (Công ty)  xin trân trọng thông báo chương trình đến Quý Khách Hàng như thông tin đính kèm,"

			pdf.text(10,32,text)
			# Loại CT

			#loai ct,loai acc
			print(pdf.get_x())
			print(pdf.get_y())
			pdf.set_y(pdf.get_y()+30)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"Loại CT",0,0,"L",1)

			pdf.cell(80,5,categoryString,0,1,"L",1)
			pdf.set_fill_color(153,204,255)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"Account","B",0,"L",1)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"{}".format(f),"B",1,"L",1)
			pdf.set_y(pdf.get_y()+5)
			#table header
			headers=['Mechanics: get/discount',"Product","Post start date","Post end date"]
			for i,header in enumerate(headers):
				if i==0:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"{}".format(headers[i]),"B",0,"L",1)
				elif i==1:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(90,5,"{}".format(headers[i]),"B",0,"L",1)
				elif i ==2:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"{}".format(headers[i]),"B",0,"L",1)
				else:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(50,5,"{}".format(headers[i]),"B",1,"L",1)

			#table data
			for ct in listloaict2:
				listMecha=excel.objects.filter(filename=file,account=f,loaiCt=ct).values("mechanicsGetORDiscount").distinct()
				for mecha in listMecha:
					datas=excel.objects.filter(filename=file,account=f,loaiCt=ct,mechanicsGetORDiscount=mecha.get("mechanicsGetORDiscount"))
					
					for row,data in enumerate(datas):
						if data.postStartDate==None or data.postEndDate==None:
							errorFlag=True
							continue
						for col,colAlphabet in enumerate(["A","B","C","D"]):
							
							mechanicsString=str(data.mechanicsGetORDiscount).replace("\n","")
							if row >0:
								mechanicsString=""	
							cellWitdhMax=80
							if pdf.get_string_width(mechanicsString) < cellWitdhMax:
								if row ==(len(datas)-1):				
									if headers[col]=='Mechanics: get/discount':
										string=mechanicsString
										pdf.cell(80,5,string,"B",0,"L")
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(90,5,"",0,0,"L")
										else:
											pdf.cell(90,5,string,0,0,"L")
									elif headers[col]=="Post start date":
										if data.postStartDate:											
											pdf.cell(60,5,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
										

									elif headers[col]=="Post end date":
										if data.postEndDate:
											pdf.cell(50,5,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
										
								else:
									if headers[col]=='Mechanics: get/discount':
										string=mechanicsString
										pdf.cell(80,5,string,0,0,"L")
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(90,5,"",0,0,"L")
										else:
											pdf.cell(90,5,string,0,0,"L")
									elif headers[col]=="Post start date":
										if data.postStartDate:
											pdf.cell(60,5,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
										
									elif headers[col]=="Post end date":
										if data.postEndDate:
											pdf.cell(50,5,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
										
							else:
								if row ==(len(datas)-1):
									mechanicsStringLen=len(mechanicsString)
									startChar=0
									maxChar=0
									holdStringEachLine=[]
									holdStringTemp=""
									while (startChar<=mechanicsStringLen):
										while (pdf.get_string_width(holdStringTemp) < cellWitdhMax) and ((startChar+maxChar)<=mechanicsStringLen):
											maxChar=maxChar+1
											holdStringTemp= mechanicsString[startChar:(maxChar+startChar)]						
										startChar=startChar+maxChar
										holdStringEachLine.append(holdStringTemp)
										#reset
										maxChar=0
										holdStringTemp=""					
									print(holdStringEachLine)
									line=len(holdStringEachLine) #Numbers of line
									if headers[col]=='Mechanics: get/discount':
										
										xPos=pdf.get_x()
										yPos=pdf.get_y()
										pdf.multi_cell(cellWitdhMax,5,mechanicsString,"B")
										pdf.set_xy(xPos+cellWitdhMax,yPos)
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(90,5*line,"",0,0,"L")
										else:
											pdf.cell(90,5*line,string,0,0,"L",)
									elif headers[col]=="Post start date":
										if data.postStartDate:
											pdf.cell(60,5*line,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
										
									elif headers[col]=="Post end date":
										if data.postEndDate:
											pdf.cell(50,5*line,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
										
								else:
									mechanicsStringLen=len(mechanicsString)
									startChar=0
									maxChar=0
									holdStringEachLine=[]
									holdStringTemp=""
									while (startChar<=mechanicsStringLen):
										while (pdf.get_string_width(holdStringTemp) < cellWitdhMax) and ((startChar+maxChar)<=mechanicsStringLen):
											maxChar=maxChar+1
											holdStringTemp= mechanicsString[startChar:(maxChar+startChar)]						
										startChar=startChar+maxChar
										holdStringEachLine.append(holdStringTemp)
										#reset
										maxChar=0
										holdStringTemp=""					
									print(holdStringEachLine)
									line=len(holdStringEachLine) #Numbers of line
									if headers[col]=='Mechanics: get/discount':
										
										xPos=pdf.get_x()
										yPos=pdf.get_y()
										pdf.multi_cell(cellWitdhMax,5,mechanicsString,0)
										pdf.set_xy(xPos+cellWitdhMax,yPos)
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(90,5*line,"",0,0,"L")
										else:
											pdf.cell(90,5*line,string,0,0,"L",)
									elif headers[col]=="Post start date":
										if data.postEndDate:
											pdf.cell(60,5*line,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
										
									elif headers[col]=="Post end date":
										if data.postEndDate:
											pdf.cell(50,5*line,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
									

			#table footer
			for i,header in enumerate(headers):
				if i==0:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"","B",0,"L",1)
				elif i==1:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(90,5,"","B",0,"L",1)
				elif i ==2:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"","B",0,"L",1)
				else:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(50,5,"","B",1,"L",1)
					
			#add footer
			pdf.set_y(pdf.get_y()+5)
			pdf.cell(100,5,"Mong Quý Khách Hàng cùng hợp tác với Công ty để đảm bảo các chương trình thực thi hiệu quả trong thời gian tới.",0,1)
			pdf.cell(100,5,"Nếu Quý Khách Hàng có bất kỳ vấn đề nào cần làm rõ, vui lòng cho KCV được biết để cùng trao đổi",0,1)

			pdf.set_y(pdf.get_y()+5)
			pdf.cell(40,5,"Trân trọng cảm ơn Quý Khách Hàng",0,1)
			pdf.cell(40,5,"Trưởng bộ phận quản lý kênh hiện đại",0,1)
			pos="{}x{}".format(round(pdf.get_x()),round(pdf.get_y()))
			print(pos) 
			
			pdf.set_y(pdf.get_y()+30)
			pdf.cell(40,5,"Phạm Nguyên Thủ",0,1)
			
			maxpageIndex=pdf.page-1
			print(maxpageIndex)
			pdf.output("pdffile.pdf")
			# pdf.write_html(html)
			annouceExist=annouceExist  +" {}_Success ".format(str(f))
			filename="THUTHONGBAO_{}_{}.pdf".format(str(f).upper(),str(datetime.now().date()))
			with open("pdffile.pdf",'rb') as pdf:
				print(1)
				pdffile=pdfFile()
				if errorFlag:
					pdffile.errorFlags=True
				pdffile.masterFile=file
				pdffile.creator=user
				pdffile.account=pdfAccount
				pdffile.loaict=categoryString
				pdffile.pos=pos
				pdffile.page_number=maxpageIndex
				pdffile.slaveFile.save(filename,File(pdf))
				if pdfEmail:			
					for email in pdfEmail:
						print(email)
						pdffile.emailExtracted.add(email)
				pdffile.save()
	elif  loaiAccount=="all" and loaict!="all":
		
		print("22222222222222222222222222222")
		listCt=loaict.split(',')
		print(listAccount)
		for f in listAccount:
			errorFlag = False
			emptyList=[]
			pdfAccount=excelAccount.objects.get(account=f)
			try:
				pdfEmail=accountEmail.objects.filter(account=pdfAccount)
			except:
				pdfEmail=[]
			for chuongTrinh in listCt:
				if excel.objects.filter(filename=file,account=f,loaiCt=chuongTrinh).exists():
					emptyList=[]
					break
				else:
					emptyList.append("{}".format(str(chuongTrinh)))

			if emptyList != []:
				annouceExist=annouceExist+" {}_no_Data ".format(str(f))
				continue
			currentAccountCates=excel.objects.filter(filename=file,account=f).values_list("loaiCt",flat=True).distinct()
			listloaict2=[]
			categoryString=""
			for ct in listCt:
				if ct in currentAccountCates:
					listloaict2.append(ct)
			print("listloaict2",listloaict2)
			
			#filter the categoryString
			
			for index,categories in enumerate(listloaict2):
				
				removeCategoryFlag=False
				#check if all data is wrong format
				if all(data is None for data in excel.objects.filter(filename=file,account=f,loaiCt=categories).values_list("postStartDate",flat=True)):
					removeCategoryFlag=True
				if all(data is None for data in excel.objects.filter(filename=file,account=f,loaiCt=categories).values_list("postEndDate",flat=True)):
					removeCategoryFlag=True
				if not removeCategoryFlag:
					if index > 0 :
						categoryString+=",{}".format(categories)
					else:
						categoryString+="{}".format(categories)
			print("categoryString",categoryString)
			print(listloaict2)
			if listloaict2 == []:
				annouceExist=annouceExist+" {}_no_Data ".format(str(f)) #empty list 2nd
				continue
			
			title = 'THÔNG BÁO VỀ CHƯƠNG TRÌNH KHUYẾN MÃI' #.encode('utf-8')
			date_year = "Tp.HCM, Ngày {}".format(str(datetime.now().date().strftime("%d/%m/%Y")))
			pdf = PDF()
			pdf.add_font('arial','',r"static/Fonts/arial.ttf",uni=True)
			pdf.add_font('arial','B',r"static/Fonts/FontsFree-Net-arial-bold.ttf",uni=True)
			pdf.add_font('arial','I',r"static/Fonts/Arial-Italic.ttf",uni=True)
			pdf.set_font('arial','', size=10)
			# pdf=FPDF("L", "mm", "A4")

			pdf.add_page("L")
			title_w = pdf.get_string_width(title) + 6

			doc_w = pdf.w
			center = (doc_w - title_w) / 2
			#tittle
			pdf.text(center,15,title)
			#logo
			pdf.image('static/image/kimberlylogo.jpg', x = 5, w = 60,h=10, y=5,type="jpg")
			# date year

			pdf.set_font('arial','', size=8)
			pdf.text(255,20,date_year)
			# Kính gửi
			pdf.set_font('arial', 'B', size=9)
			text = 'Kính gửi: Quý Khách Hàng Kênh Hiện Đại'

			pdf.text(10, 25, text)
			# Tên chương trình
			pdf.set_font('arial','', size=8)
			text = "Công ty TNHH Kimberly-Clark Việt Nam (Công ty)  xin trân trọng thông báo chương trình đến Quý Khách Hàng như thông tin đính kèm,"

			pdf.text(10,32,text)
			# Loại CT

			#loai ct,loai acc
			print(pdf.get_x())
			print(pdf.get_y())
			pdf.set_y(pdf.get_y()+30)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"Loại CT",0,0,"L",1)

			pdf.cell(80,5,categoryString,0,1,"L",1)
			pdf.set_fill_color(153,204,255)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"Account","B",0,"L",1)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"{}".format(f),"B",1,"L",1)
			pdf.set_y(pdf.get_y()+5)
			#table header
			headers=['Mechanics: get/discount',"Product","Post start date","Post end date"]
			for i,header in enumerate(headers):
				if i==0:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"{}".format(headers[i]),"B",0,"L",1)
				elif i==1:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(90,5,"{}".format(headers[i]),"B",0,"L",1)
				elif i ==2:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"{}".format(headers[i]),"B",0,"L",1)
				else:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(50,5,"{}".format(headers[i]),"B",1,"L",1)
			
			#table data
			for ct in listloaict2:
				listMecha=excel.objects.filter(filename=file,account=f,loaiCt=ct).values("mechanicsGetORDiscount").distinct()
				for mecha in listMecha:
					datas=excel.objects.filter(filename=file,account=f,loaiCt=ct,mechanicsGetORDiscount=mecha.get("mechanicsGetORDiscount"))
					print(datas)
					for row,data in enumerate(datas):
						if data.postStartDate==None or data.postEndDate==None:
							errorFlag=True
							continue
						for col,colAlphabet in enumerate(["A","B","C","D"]):
							mechanicsString=str(data.mechanicsGetORDiscount).replace("\n","")
							if row >0:
								mechanicsString=""	
							cellWitdhMax=80
							if pdf.get_string_width(mechanicsString) < cellWitdhMax:
								if row ==(len(datas)-1):				
									if headers[col]=='Mechanics: get/discount':
										string=mechanicsString
										pdf.cell(80,5,string,"B",0,"L")
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(90,5,"",0,0,"L")
										else:
											pdf.cell(90,5,string,0,0,"L")
									elif headers[col]=="Post start date":
										if data.postStartDate:
											pdf.cell(60,5,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
										
									elif headers[col]=="Post end date":
										if data.postEndDate:
											pdf.cell(50,5,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
										
								else:
									if headers[col]=='Mechanics: get/discount':
										string=mechanicsString
										pdf.cell(80,5,string,0,0,"L")
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(90,5,"",0,0,"L")
										else:
											pdf.cell(90,5,string,0,0,"L")
									elif headers[col]=="Post start date":
										if data.postStartDate:
											pdf.cell(60,5,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
										
									elif headers[col]=="Post end date":
										if data.postEndDate:
											pdf.cell(50,5,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
										
							else:
								if row ==(len(datas)-1):
									mechanicsStringLen=len(mechanicsString)
									startChar=0
									maxChar=0
									holdStringEachLine=[]
									holdStringTemp=""
									while (startChar<=mechanicsStringLen):
										while (pdf.get_string_width(holdStringTemp) < cellWitdhMax) and ((startChar+maxChar)<=mechanicsStringLen):
											maxChar=maxChar+1
											holdStringTemp= mechanicsString[startChar:(maxChar+startChar)]						
										startChar=startChar+maxChar
										holdStringEachLine.append(holdStringTemp)
										#reset
										maxChar=0
										holdStringTemp=""					
									print(holdStringEachLine)
									line=len(holdStringEachLine) #Numbers of line
									if headers[col]=='Mechanics: get/discount':
										
										xPos=pdf.get_x()
										yPos=pdf.get_y()
										pdf.multi_cell(cellWitdhMax,5,mechanicsString,"B")
										pdf.set_xy(xPos+cellWitdhMax,yPos)
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(90,5*line,"",0,0,"L")
										else:
											pdf.cell(90,5*line,string,0,0,"L",)
									elif headers[col]=="Post start date":
										if data.postStartDate:
											pdf.cell(60,5*line,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
										
									elif headers[col]=="Post end date":
										if data.postEndDate:
											pdf.cell(50,5*line,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
										
								else:
									mechanicsStringLen=len(mechanicsString)
									startChar=0
									maxChar=0
									holdStringEachLine=[]
									holdStringTemp=""
									while (startChar<=mechanicsStringLen):
										while (pdf.get_string_width(holdStringTemp) < cellWitdhMax) and ((startChar+maxChar)<=mechanicsStringLen):
											maxChar=maxChar+1
											holdStringTemp= mechanicsString[startChar:(maxChar+startChar)]						
										startChar=startChar+maxChar
										holdStringEachLine.append(holdStringTemp)
										#reset
										maxChar=0
										holdStringTemp=""					
									print(holdStringEachLine)
									line=len(holdStringEachLine) #Numbers of line
									if headers[col]=='Mechanics: get/discount':
										
										xPos=pdf.get_x()
										yPos=pdf.get_y()
										pdf.multi_cell(cellWitdhMax,5,mechanicsString,0)
										pdf.set_xy(xPos+cellWitdhMax,yPos)
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(90,5*line,"",0,0,"L")
										else:
											pdf.cell(90,5*line,string,0,0,"L",)
									elif headers[col]=="Post start date":
										if data.postStartDate:
											pdf.cell(60,5*line,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
										
									elif headers[col]=="Post end date":
										if data.postEndDate:
											pdf.cell(50,5*line,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
										

			#table footer
			for i,header in enumerate(headers):
				if i==0:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"","B",0,"L",1)
				elif i==1:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(90,5,"","B",0,"L",1)
				elif i ==2:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"","B",0,"L",1)
				else:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(50,5,"","B",1,"L",1)
					
			#add footer
			pdf.set_y(pdf.get_y()+5)
			pdf.cell(100,5,"Mong Quý Khách Hàng cùng hợp tác với Công ty để đảm bảo các chương trình thực thi hiệu quả trong thời gian tới.",0,1)
			pdf.cell(100,5,"Nếu Quý Khách Hàng có bất kỳ vấn đề nào cần làm rõ, vui lòng cho KCV được biết để cùng trao đổi",0,1)

			pdf.set_y(pdf.get_y()+5)
			pdf.cell(40,5,"Trân trọng cảm ơn Quý Khách Hàng",0,1)
			pdf.cell(40,5,"Trưởng bộ phận quản lý kênh hiện đại",0,1)
			pos="{}x{}".format(round(pdf.get_x()),round(pdf.get_y()))
			print(pos)
			
			pdf.set_y(pdf.get_y()+30)
			pdf.cell(40,5,"Phạm Nguyên Thủ",0,1)
			
			
			pdf.set_y(pdf.get_y()+5)
		
			
			maxpageIndex=pdf.page-1
			print(maxpageIndex)
			pdf.output("pdffile.pdf")
			# pdf.write_html(html)
			
			
			
			annouceExist=annouceExist  +" {}_Success ".format(str(f))
			filename="THUTHONGBAO_{}_{}.pdf".format(str(f).upper(),str(datetime.now().date()))
			
			print(filename)
			with open("pdffile.pdf",'rb') as pdf:
				print(1)
				pdffile=pdfFile()
				pdffile.masterFile=file
				if errorFlag:
					pdffile.errorFlags=True
				pdffile.creator=user
				pdffile.account=pdfAccount
				pdffile.loaict=categoryString
				pdffile.pos=pos
				pdffile.page_number=maxpageIndex
				pdffile.slaveFile.save(filename,File(pdf))
				if pdfEmail:			
					for email in pdfEmail:
						print(email)
						pdffile.emailExtracted.add(email)
				pdffile.save()
	elif loaiAccount!="all" and loaict=="all":
		print("#33333333333333333333333")
		listAcc=loaiAccount.split(',')	
	# 	print(listAcc)
		for f in listAcc:
			errorFlag = False
			emptyList=[]
			pdfAccount=excelAccount.objects.get(account=f)
			try:
				pdfEmail=accountEmail.objects.filter(account=pdfAccount)
			except:
				pdfEmail=[]
			for chuongTrinh in listloaict:
				if excel.objects.filter(filename=file,account=f,loaiCt=chuongTrinh).exists():
					emptyList=[]
					break
				else:
					emptyList.append("{}{}".format(f,str(chuongTrinh)))

			if emptyList != []:
				annouceExist=annouceExist+" {}_no_Data ".format(str(f))
				continue
			currentAccountCates=excel.objects.filter(filename=file,account=f).values("loaiCt").distinct()
			listloaict2=[]
			categoryString=""
			for i,b in enumerate(currentAccountCates):
				listloaict2.append(b['loaiCt'])
			#filter the categoryString
			for index,categories in enumerate(currentAccountCates):
				removeCategoryFlag=False
				#check if all data is wrong format
				if all(data is None for data in excel.objects.filter(filename=file,account=f,loaiCt=categories['loaiCt']).values_list("postStartDate",flat=True)):
					removeCategoryFlag=True
				if all(data is None for data in excel.objects.filter(filename=file,account=f,loaiCt=categories['loaiCt']).values_list("postEndDate",flat=True)):
					removeCategoryFlag=True
				if removeCategoryFlag:
					listloaict2.remove("{}".format(str(categories['loaiCt'])))
				else:
					if index > 0 :
						categoryString+=",{}".format(categories['loaiCt'])
					else:
						categoryString+="{}".format(categories['loaiCt'])
			if listloaict2 == []:
				annouceExist=annouceExist+" {}_no_Data ".format(str(f)) #empty list 2nd
				continue
			title = 'THÔNG BÁO VỀ CHƯƠNG TRÌNH KHUYẾN MÃI' #.encode('utf-8')
			date_year = "Tp.HCM, Ngày {}".format(str(datetime.now().date().strftime("%d/%m/%Y")))
			pdf = PDF()
			pdf.add_font('arial','',r"static/Fonts/arial.ttf",uni=True)
			pdf.add_font('arial','B',r"static/Fonts/FontsFree-Net-arial-bold.ttf",uni=True)
			pdf.add_font('arial','I',r"static/Fonts/Arial-Italic.ttf",uni=True)
			pdf.set_font('arial','', size=10)
			# pdf=FPDF("L", "mm", "A4")

			pdf.add_page("L")
			title_w = pdf.get_string_width(title) + 6

			doc_w = pdf.w
			center = (doc_w - title_w) / 2
			#tittle
			pdf.text(center,15,title)
			#logo
			pdf.image('static/image/kimberlylogo.jpg', x = 5, w = 60,h=10, y=5,type="jpg")
			# date year

			pdf.set_font('arial','', size=8)
			pdf.text(255,20,date_year)
			# Kính gửi
			pdf.set_font('arial', 'B', size=9)
			text = 'Kính gửi: Quý Khách Hàng Kênh Hiện Đại'

			pdf.text(10, 25, text)
			# Tên chương trình
			pdf.set_font('arial','', size=8)
			text = "Công ty TNHH Kimberly-Clark Việt Nam (Công ty)  xin trân trọng thông báo chương trình đến Quý Khách Hàng như thông tin đính kèm,"

			pdf.text(10,32,text)
			# Loại CT

			#loai ct,loai acc
			print(pdf.get_x())
			print(pdf.get_y())
			pdf.set_y(pdf.get_y()+30)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"Loại CT",0,0,"L",1)

			pdf.cell(80,5,categoryString,0,1,"L",1)
			pdf.set_fill_color(153,204,255)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"Account","B",0,"L",1)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"{}".format(f),"B",1,"L",1)
			pdf.set_y(pdf.get_y()+5)
	# 		#table header
			headers=['Mechanics: get/discount',"Product","Post start date","Post end date"]
			for i,header in enumerate(headers):
				if i==0:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"{}".format(headers[i]),"B",0,"L",1)
				elif i==1:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(90,5,"{}".format(headers[i]),"B",0,"L",1)
				elif i ==2:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"{}".format(headers[i]),"B",0,"L",1)
				else:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(50,5,"{}".format(headers[i]),"B",1,"L",1)

	# 		#table data
			for ct in listloaict2:
				listMecha=excel.objects.filter(filename=file,account=f,loaiCt=ct).values("mechanicsGetORDiscount").distinct()
				for mecha in listMecha:
					datas=excel.objects.filter(filename=file,account=f,loaiCt=ct,mechanicsGetORDiscount=mecha.get("mechanicsGetORDiscount"))
					print(datas)
					for row,data in enumerate(datas):
						if data.postStartDate==None or data.postEndDate==None:
							errorFlag=True
							continue
						for col,colAlphabet in enumerate(["A","B","C","D"]):
							
							mechanicsString=str(data.mechanicsGetORDiscount).replace("\n","")
							if row >0:
								mechanicsString=""	
							cellWitdhMax=80
							if pdf.get_string_width(mechanicsString) < cellWitdhMax:
								if row ==(len(datas)-1):				
									if headers[col]=='Mechanics: get/discount':
										string=mechanicsString
										pdf.cell(80,5,string,"B",0,"L")
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(90,5,"",0,0,"L")
										else:
											pdf.cell(90,5,string,0,0,"L")
									elif headers[col]=="Post start date":
										if data.postStartDate:
											pdf.cell(60,5,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")

									elif headers[col]=="Post end date":
										if data.postEndDate:
											pdf.cell(50,5,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")

								else:
									if headers[col]=='Mechanics: get/discount':
										string=mechanicsString
										pdf.cell(80,5,string,0,0,"L")
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(90,5,"",0,0,"L")
										else:
											pdf.cell(90,5,string,0,0,"L")
									elif headers[col]=="Post start date":
										if data.postStartDate:
											pdf.cell(60,5,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")

									elif headers[col]=="Post end date":
										if data.postEndDate:
											pdf.cell(50,5,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")

							else:
								if row ==(len(datas)-1):
									mechanicsStringLen=len(mechanicsString)
									startChar=0
									maxChar=0
									holdStringEachLine=[]
									holdStringTemp=""
									while (startChar<=mechanicsStringLen):
										while (pdf.get_string_width(holdStringTemp) < cellWitdhMax) and ((startChar+maxChar)<=mechanicsStringLen):
											maxChar=maxChar+1
											holdStringTemp= mechanicsString[startChar:(maxChar+startChar)]						
										startChar=startChar+maxChar
										holdStringEachLine.append(holdStringTemp)
										#reset
										maxChar=0
										holdStringTemp=""					
									print(holdStringEachLine)
									line=len(holdStringEachLine) #Numbers of line
									if headers[col]=='Mechanics: get/discount':
										
										xPos=pdf.get_x()
										yPos=pdf.get_y()
										pdf.multi_cell(cellWitdhMax,5,mechanicsString,"B")
										pdf.set_xy(xPos+cellWitdhMax,yPos)
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(90,5*line,"",0,0,"L")
										else:
											pdf.cell(90,5*line,string,0,0,"L",)
									elif headers[col]=="Post start date":
										if data.postStartDate:
											pdf.cell(60,5*line,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
	
									elif headers[col]=="Post end date":
										if data.postEndDate:
											pdf.cell(50,5*line,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")

								else:
									mechanicsStringLen=len(mechanicsString)
									startChar=0
									maxChar=0
									holdStringEachLine=[]
									holdStringTemp=""
									while (startChar<=mechanicsStringLen):
										while (pdf.get_string_width(holdStringTemp) < cellWitdhMax) and ((startChar+maxChar)<=mechanicsStringLen):
											maxChar=maxChar+1
											holdStringTemp= mechanicsString[startChar:(maxChar+startChar)]						
										startChar=startChar+maxChar
										holdStringEachLine.append(holdStringTemp)
										#reset
										maxChar=0
										holdStringTemp=""					
									print(holdStringEachLine)
									line=len(holdStringEachLine) #Numbers of line
									if headers[col]=='Mechanics: get/discount':
										
										xPos=pdf.get_x()
										yPos=pdf.get_y()
										pdf.multi_cell(cellWitdhMax,5,mechanicsString,0)
										pdf.set_xy(xPos+cellWitdhMax,yPos)
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(90,5*line,"",0,0,"L")
										else:
											pdf.cell(90,5*line,string,0,0,"L",)
									elif headers[col]=="Post start date":
										if data.postStartDate:
											pdf.cell(60,5*line,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")

									elif headers[col]=="Post end date":
										if data.postEndDate:
											pdf.cell(50,5*line,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
										

			#table footer
			for i,header in enumerate(headers):
				if i==0:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"","B",0,"L",1)
				elif i==1:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(90,5,"","B",0,"L",1)
				elif i ==2:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"","B",0,"L",1)
				else:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(50,5,"","B",1,"L",1)
					
			#add footer
			pdf.set_y(pdf.get_y()+5)
			pdf.cell(100,5,"Mong Quý Khách Hàng cùng hợp tác với Công ty để đảm bảo các chương trình thực thi hiệu quả trong thời gian tới.",0,1)
			pdf.cell(100,5,"Nếu Quý Khách Hàng có bất kỳ vấn đề nào cần làm rõ, vui lòng cho KCV được biết để cùng trao đổi",0,1)

			pdf.set_y(pdf.get_y()+5)
			pdf.cell(40,5,"Trân trọng cảm ơn Quý Khách Hàng",0,1)
			pdf.cell(40,5,"Trưởng bộ phận quản lý kênh hiện đại",0,1)
			pos="{}x{}".format(round(pdf.get_x()),round(pdf.get_y()))
			print(pos)
			
			pdf.set_y(pdf.get_y()+30)
			pdf.cell(40,5,"Phạm Nguyên Thủ",0,1)
			
			
			
			print(pos)
			maxpageIndex=pdf.page-1
			print(maxpageIndex)
			pdf.output("pdffile.pdf")
			# pdf.write_html(html)
			
			
			annouceExist=annouceExist  +" {}_Success ".format(str(f))
			filename="THUTHONGBAO_{}_{}.pdf".format(str(f).upper(),str(datetime.now().date()))		
			print(filename)
			with open("pdffile.pdf",'rb') as pdf:
				print(1)
				pdffile=pdfFile()
				if errorFlag:
					pdffile.errorFlags=True
				pdffile.masterFile=file
				pdffile.creator=user
				pdffile.account=pdfAccount
				pdffile.loaict=categoryString
				pdffile.pos=pos
				pdffile.page_number=maxpageIndex
				pdffile.slaveFile.save(filename,File(pdf))
				if pdfEmail:			
					for email in pdfEmail:
						print(email)
						pdffile.emailExtracted.add(email)
				pdffile.save()
	elif loaiAccount!="all" and loaict!="all":
	# 	print("4")
		listAcc=loaiAccount.split(',')
		listCt=loaict.split(',')
		
		for f in listAcc:
			errorFlag=False
			emptyList=[]
			pdfAccount=excelAccount.objects.get(account=f)
			try:
				pdfEmail=accountEmail.objects.filter(account=pdfAccount)
			except:
				pdfEmail=[]
			for chuongTrinh in listCt:
				if excel.objects.filter(filename=file,account=f,loaiCt=chuongTrinh).exists():
					emptyList=[]
					break
				else:
					emptyList.append("{}".format(str(chuongTrinh)))

			if emptyList != []:
				annouceExist=annouceExist+" {}_no_Data ".format(str(f))
				continue
			currentAccountCates=excel.objects.filter(filename=file,account=f).values_list("loaiCt",flat=True).distinct()
			listloaict2=[]
			categoryString=""
			for ct in listCt:
				if ct in currentAccountCates:
					listloaict2.append(ct)
			print("listloaict2",listloaict2)
			
			#filter the categoryString
			
			for index,categories in enumerate(listloaict2):
				
				removeCategoryFlag=False
				#check if all data is wrong format
				if all(data is None for data in excel.objects.filter(filename=file,account=f,loaiCt=categories).values_list("postStartDate",flat=True)):
					removeCategoryFlag=True
				if all(data is None for data in excel.objects.filter(filename=file,account=f,loaiCt=categories).values_list("postEndDate",flat=True)):
					removeCategoryFlag=True
				if not removeCategoryFlag:
					if index > 0 :
						categoryString+=",{}".format(categories)
					else:
						categoryString+="{}".format(categories)
			print("categoryString",categoryString)
			print(listloaict2)
			if listloaict2 == []:
				annouceExist=annouceExist+" {}_no_Data ".format(str(f)) #empty list 2nd
				continue
			
			title = 'THÔNG BÁO VỀ CHƯƠNG TRÌNH KHUYẾN MÃI' #.encode('utf-8')
			date_year = "Tp.HCM, Ngày {}".format(str(datetime.now().date().strftime("%d/%m/%Y")))
			pdf = PDF()
			pdf.add_font('arial','',r"static/Fonts/arial.ttf",uni=True)
			pdf.add_font('arial','B',r"static/Fonts/FontsFree-Net-arial-bold.ttf",uni=True)
			pdf.add_font('arial','I',r"static/Fonts/Arial-Italic.ttf",uni=True)
			pdf.set_font('arial','', size=10)
			# pdf=FPDF("L", "mm", "A4")

			pdf.add_page("L")
			title_w = pdf.get_string_width(title) + 6

			doc_w = pdf.w
			center = (doc_w - title_w) / 2
			#tittle
			pdf.text(center,15,title)
			#logo
			pdf.image('static/image/kimberlylogo.jpg', x = 5, w = 60,h=10, y=5,type="jpg")
			# date year

			pdf.set_font('arial','', size=8)
			pdf.text(255,20,date_year)
			# Kính gửi
			pdf.set_font('arial', 'B', size=9)
			text = 'Kính gửi: Quý Khách Hàng Kênh Hiện Đại'

			pdf.text(10, 25, text)
			# Tên chương trình
			pdf.set_font('arial','', size=8)
			text = "Công ty TNHH Kimberly-Clark Việt Nam (Công ty)  xin trân trọng thông báo chương trình đến Quý Khách Hàng như thông tin đính kèm,"

			pdf.text(10,32,text)
			# Loại CT

			#loai ct,loai acc
			print(pdf.get_x())
			print(pdf.get_y())
			pdf.set_y(pdf.get_y()+30)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"Loại CT",0,0,"L",1)

			pdf.cell(80,5,categoryString,0,1,"L",1)
			pdf.set_fill_color(153,204,255)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"Account","B",0,"L",1)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"{}".format(f),"B",1,"L",1)
			pdf.set_y(pdf.get_y()+5)
	# 		#table header
			headers=['Mechanics: get/discount',"Product","Post start date","Post end date"]
			for i,header in enumerate(headers):
				if i==0:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"{}".format(headers[i]),"B",0,"L",1)
				elif i==1:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(90,5,"{}".format(headers[i]),"B",0,"L",1)
				elif i ==2:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"{}".format(headers[i]),"B",0,"L",1)
				else:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(50,5,"{}".format(headers[i]),"B",1,"L",1)

			#table data
			for ct in listloaict2:
				listMecha=excel.objects.filter(filename=file,account=f,loaiCt=ct).values("mechanicsGetORDiscount").distinct()
				for mecha in listMecha:
					datas=excel.objects.filter(filename=file,account=f,loaiCt=ct,mechanicsGetORDiscount=mecha.get("mechanicsGetORDiscount"))
					print(datas)
					for row,data in enumerate(datas):
						if data.postStartDate==None or data.postEndDate==None:
								errorFlag=True
								continue
						for col,colAlphabet in enumerate(["A","B","C","D"]):
							
							mechanicsString=str(data.mechanicsGetORDiscount).replace("\n","")
							if row >0:
								mechanicsString=""	
							cellWitdhMax=80
							if pdf.get_string_width(mechanicsString) < cellWitdhMax:
								if row ==(len(datas)-1):				
									if headers[col]=='Mechanics: get/discount':
										string=mechanicsString
										pdf.cell(80,5,string,"B",0,"L")
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(90,5,"",0,0,"L")
										else:
											pdf.cell(90,5,string,0,0,"L")
									elif headers[col]=="Post start date":
										if data.postStartDate:
											pdf.cell(60,5,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
										
									elif headers[col]=="Post end date":
										if data.postEndDate:
											pdf.cell(50,5,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
										
								else:
									if headers[col]=='Mechanics: get/discount':
										string=mechanicsString
										pdf.cell(80,5,string,0,0,"L")
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(90,5,"",0,0,"L")
										else:
											pdf.cell(90,5,string,0,0,"L")
									elif headers[col]=="Post start date":
										if data.postStartDate:
											pdf.cell(60,5,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
										
									elif headers[col]=="Post end date":
										if data.postEndDate:
											pdf.cell(50,5,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
										
							else:
								if row ==(len(datas)-1):
									mechanicsStringLen=len(mechanicsString)
									startChar=0
									maxChar=0
									holdStringEachLine=[]
									holdStringTemp=""
									while (startChar<=mechanicsStringLen):
										while (pdf.get_string_width(holdStringTemp) < cellWitdhMax) and ((startChar+maxChar)<=mechanicsStringLen):
											maxChar=maxChar+1
											holdStringTemp= mechanicsString[startChar:(maxChar+startChar)]						
										startChar=startChar+maxChar
										holdStringEachLine.append(holdStringTemp)
										#reset
										maxChar=0
										holdStringTemp=""					
									print(holdStringEachLine)
									line=len(holdStringEachLine) #Numbers of line
									if headers[col]=='Mechanics: get/discount':
										
										xPos=pdf.get_x()
										yPos=pdf.get_y()
										pdf.multi_cell(cellWitdhMax,5,mechanicsString,"B")
										pdf.set_xy(xPos+cellWitdhMax,yPos)
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(90,5*line,"",0,0,"L")
										else:
											pdf.cell(90,5*line,string,0,0,"L",)
									elif headers[col]=="Post start date":
										if data.postStartDate:
											pdf.cell(60,5*line,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
										
									elif headers[col]=="Post end date":
										if data.postEndDate:
											pdf.cell(50,5*line,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
										
								else:
									mechanicsStringLen=len(mechanicsString)
									startChar=0
									maxChar=0
									holdStringEachLine=[]
									holdStringTemp=""
									while (startChar<=mechanicsStringLen):
										while (pdf.get_string_width(holdStringTemp) < cellWitdhMax) and ((startChar+maxChar)<=mechanicsStringLen):
											maxChar=maxChar+1
											holdStringTemp= mechanicsString[startChar:(maxChar+startChar)]						
										startChar=startChar+maxChar
										holdStringEachLine.append(holdStringTemp)
										#reset
										maxChar=0
										holdStringTemp=""					
									print(holdStringEachLine)
									line=len(holdStringEachLine) #Numbers of line
									if headers[col]=='Mechanics: get/discount':
										
										xPos=pdf.get_x()
										yPos=pdf.get_y()
										pdf.multi_cell(cellWitdhMax,5,mechanicsString,0)
										pdf.set_xy(xPos+cellWitdhMax,yPos)
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(90,5*line,"",0,0,"L")
										else:
											pdf.cell(90,5*line,string,0,0,"L",)
									elif headers[col]=="Post start date":
										if data.postStartDate:
											pdf.cell(60,5*line,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
										
									elif headers[col]=="Post end date":
										if data.postEndDate:
											pdf.cell(50,5*line,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
										

			#table footer
			for i,header in enumerate(headers):
				if i==0:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"","B",0,"L",1)
				elif i==1:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(90,5,"","B",0,"L",1)
				elif i ==2:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"","B",0,"L",1)
				else:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(50,5,"","B",1,"L",1)
					
			#add footer
			pdf.set_y(pdf.get_y()+5)
			pdf.cell(100,5,"Mong Quý Khách Hàng cùng hợp tác với Công ty để đảm bảo các chương trình thực thi hiệu quả trong thời gian tới.",0,1)
			pdf.cell(100,5,"Nếu Quý Khách Hàng có bất kỳ vấn đề nào cần làm rõ, vui lòng cho KCV được biết để cùng trao đổi",0,1)

			pdf.set_y(pdf.get_y()+5)
			pdf.cell(40,5,"Trân trọng cảm ơn Quý Khách Hàng",0,1)
			pdf.cell(40,5,"Trưởng bộ phận quản lý kênh hiện đại",0,1)
			pos="{}x{}".format(round(pdf.get_x()),round(pdf.get_y()))
			print(pos)
			
			pdf.set_y(pdf.get_y()+30)
			pdf.cell(40,5,"Phạm Nguyên Thủ",0,1)
			
			pdf.set_y(pdf.get_y()+5)
		
			
			maxpageIndex=pdf.page-1
			print(maxpageIndex)
			pdf.output("pdffile.pdf")
			# pdf.write_html(html)
			
			
			annouceExist=annouceExist  +" {}_Success ".format(str(f))
			filename="THUTHONGBAO_{}_{}.pdf".format(str(f).upper(),str(datetime.now().date()))	
			print(filename)
			with open("pdffile.pdf",'rb') as pdf:
				print(1)
				pdffile=pdfFile()
				if errorFlag:
					pdffile.errorFlags=True
				pdffile.masterFile=file
				pdffile.creator=user
				pdffile.account=pdfAccount
				pdffile.loaict=categoryString
				pdffile.pos=pos
				pdffile.page_number=maxpageIndex
				pdffile.slaveFile.save(filename,File(pdf))
				if pdfEmail:			
					for email in pdfEmail:
						print(email)
						pdffile.emailExtracted.add(email)
				pdffile.save()
	return annouceExist


