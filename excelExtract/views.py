from audioop import reverse
from urllib import response
from django.shortcuts import render,redirect
import sys
from rest_framework.decorators import api_view
from excelExtract.forms import uploadDocumentForm
from excelExtract.models import document, excelAccount,pdfFile, excel
from . import excelExtract
from user.models import Profile
from django.db import transaction
from rest_framework import status
from django.conf import settings
from django.http import HttpResponse
from rest_framework.response import Response
from django.views.decorators.clickjacking import xframe_options_sameorigin
from utils import send_email
from datetime import datetime
import requests
import json
from django.core.files import File
from django.db.models import Q
import fitz
from django.urls import reverse_lazy
import openpyxl
from openpyxl.styles import Alignment,NamedStyle
from openpyxl.writer.excel import save_virtual_workbook
# Create your views here.
def detect_position(pdf_file_location):
	pdf = fitz.open(pdf_file_location)
	
	page_0 = pdf.load_page(0)
	page_width, page_height = page_0.rect.width, page_0.rect.height
	y1 = 200
	marked_page_num=0
	x0 = 200
	
	search_text = 'Trưởng bộ phận quản lý kênh hiện đại'
	for i in range(pdf.page_count):
		text_instances=pdf.load_page(i).search_for(search_text)
		if  text_instances != []:
			marked_page_num = i
			x0=text_instances[0].x0
			y1=text_instances[0].y1
			
	if y1 >  595:
		marked_page_num = marked_page_num + 1
		left = x0
		bottom = 10.5 * 3
	else:
		left = x0
		bottom = y1 

	# convert to milimeter  
	left = left
	bottom = (page_height - bottom) - 10.5*2 - 50  # 10.5 la chieu cao 1 dong dong, 50 trong 50x50

	return left, bottom, marked_page_num
@xframe_options_sameorigin
def kcToolPage(request):
	if request.user.is_authenticated:
		form=uploadDocumentForm()
		files=document.objects.all()
		pdffiles=pdfFile.objects.all().order_by("-id")
		demoPdfFiles=pdfFile.objects.last()
		numberUnsignepdfs=len(pdfFile.objects.filter(confirmed=False))
		if request.method=='POST':
			with transaction.atomic():
				try:
					file=request.FILES['document']
					print(file)
					if file:
						try:
							user = request.user
							profile = Profile.objects.get(user=user)
							excelExtract.importDataExcel(file, user=profile)
						except:
							excelExtract.importDataExcel(file)
				except:
					pass
		return render(request,"KCtool/KCtool.html",{"form":form,"files":files,"pdffiles":pdffiles,"demoPdfFiles":demoPdfFiles,"numberUnsignepdfs":numberUnsignepdfs, "active_id":1})
	else :
		return HttpResponse("not authen")


def waitConfirmDoc(request):
	if request.user.is_authenticated:
		user=request.user
		numberunconfirmpdfs=len(pdfFile.objects.filter(confirmed=False))
		unconfirmpdfs=pdfFile.objects.filter(confirmed=False).order_by("-id")
		accountList=excelAccount.objects.all()
		profile=Profile.objects.get(user=user)
		listaccount=excelAccount.objects.filter(responsibleBy=profile)
		if request.GET.get("account",None):
			account=excelAccount.objects.filter(account=request.GET.get("account"))[0]
			print(account.pk)
			unconfirmpdfs = pdfFile.objects.filter(confirmed=False,account=account).order_by("-id")
		return render(request,"KCtool/waitingsigndoc.html",{"numberunconfirmpdfs":numberunconfirmpdfs,"unconfirmpdfs":unconfirmpdfs, "URL":settings.URL, "active_id":2,"accountList":accountList,"user":user,"listaccount":listaccount,"account":request.GET.get("account",None)})
	else :
		return HttpResponse("not authen")
def signedDoc(request):
	if request.user.is_authenticated:
		user=request.user
		numbersendedpdfs=len(pdfFile.objects.filter(sended=True))
		pdfs=pdfFile.objects.filter(confirmed=True).order_by("sended")
		accountList=excelAccount.objects.all()
		if request.GET.get("key_word",""):
			key_word = request.GET.get("key_word")
			pdfs=pdfs.filter(slaveFile__icontains=key_word).order_by("sended")	
		if request.GET.get("account",None):
			account=excelAccount.objects.filter(account=request.GET.get("account"))[0]
			print(account.pk)			
			pdfs = pdfs.filter(account=account).order_by("sended")
		if request.GET.get("todate",None):
			todate=request.GET.get("todate")
			pdfs = pdfs.filter(createdTime__date__lte=todate).order_by("sended")|pdfs.filter(sendingTime__date__lte=todate).order_by("sended")
		if request.GET.get("fromdate",None):	
			fromdate=request.GET.get("fromdate")
			pdfs = pdfs.filter(createdTime__date__gte=fromdate).order_by("sended")|pdfs.filter(sendingTime__date__gte=fromdate).order_by("sended")
		if request.GET.get("fromdate2",None) and request.GET.get("fromdate2",None):	
			fromdate2=request.GET.get("fromdate2")
			todate2=request.GET.get("todate2")
			# actlogs = pdfFile.objects.filter(createdTime__date__gte=fromdate2,createdTime__date__lte=todate2)|pdfFile.objects.filter(sendingTime__date__gte=fromdate2,createdTime__date__lte=todate2)
			return export_hnk_ticket_excel(fromdate2,todate2)
			
			
		return render(request,"KCtool/signedDoc.html",{"numbersendedpdfs":numbersendedpdfs,"pdfs":pdfs, "active_id":3,"user":user,"accountList":accountList,"key_word":request.GET.get("key_word",""),"account":request.GET.get("account",None),"fromdate":request.GET.get("fromdate"),"todate":request.GET.get("todate"),"fromdate2":request.GET.get("fromdate2"),"todate2":request.GET.get("todate2")})

	else :
		return HttpResponse("not authen")
# def excelToListPdfs(request):  
#             return Response(request,"KCtool/KCTool.html")
@api_view(["POST"])
def getIdList(request):
	if request.method=='POST':
		print(request.data)
		# loaict=request.data['loaict']
		# listId=request.POST.getlist('file')
		# for  f in listId:
		#     loaiAccount=request.data['fileID{}'.format(f)]
		#     excelExtract.exportFiles(loaict=loaict,fileID=f,loaiAccount=loaiAccount)  
		return redirect("KCTool:kcToolPage")
	

@api_view(["POST"])
def create_pdf(request):
	try:
		with transaction.atomic():
			print(request.data)
			id_excel = int(request.data["pk_excel"])
			values_category = request.data["values_category"]
			values_account = request.data["values_account"]
			print(values_account)
			print(values_category)
			user=request.user
			profile = Profile.objects.get(user=user)
			annouce=excelExtract.exportFiles(loaict=values_category,fileID=id_excel,loaiAccount=values_account,user=profile) 
			print(annouce)
			return Response({"annouce":annouce}, status=status.HTTP_200_OK)
	except:
		return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
def getListAccount(request):
	# try:
		arr_listaccounts=[]
		arr_loaiCt = []
		list_accounts = ""
		list_loaiCt = ""
		pk = request.data["pk_excel"]
		file=document.objects.get(pk=int(pk))
		list_accounts +=f'''
					<option group="list_accounts_group" value="all" data-badge="">All</option>
				'''
		list_loaiCt+= f'''
					<option group="list_LoaiCT_group" value="all" data-badge="">All</option>
				'''
		for f in excel.objects.filter(filename=file):
			if f.account not in arr_listaccounts:

				arr_listaccounts.append(f.account)
				list_accounts += f'''
					<option group="list_accounts_group" value="{f.account}" data-badge="">{f.account}</option>
				'''
			if f.loaiCt not in arr_loaiCt:
				arr_loaiCt.append(f.loaiCt)
				list_loaiCt += f'''
					<option group="list_LoaiCT_group" value="{f.loaiCt}" data-badge="">{f.loaiCt}</option>
				'''
		
		return Response({"Status":"Success", "list_accounts": list_accounts, "list_loaiCt":list_loaiCt},  status=status.HTTP_200_OK)
	# except:
	# 	return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(["POST"])
def sign_pdf(request):
	try:
		
		list_id_pdf_file = request.data["list_id_pdf_file"]
		list_id=list_id_pdf_file.split(",")
		for i in list_id:
			pdf = pdfFile.objects.get(pk=int(i))
			pdf.confirmed = True
			pdf.save()
		print(list_id_pdf_file)
		numberunsignepdfs=len(pdfFile.objects.filter(signed=False,confirmed=True))
		send_email.send_noti_to_partner_sign_by_email(["{} văn bản cần kí".format(str(numberunsignepdfs))], "khang.huynhquoc@kcc.com")
		return Response({"code":"00"}, status=status.HTTP_200_OK)
	except:
		err_mess = sys.exc_info()[0].__name__ + ": "+ str(sys.exc_info()[1])
		print(err_mess)
		return Response({"err":err_mess},status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(["POST"])
def send_pdf(request):
	account_data={"user_name": "baynguyen2000@gmail.com",
  				"password": "Tu12345@"}
	headers = { 
	'Content-Type':'application/json' 
   }
	response_obj = requests.post(r"https://api-testing.pvs.com.vn/user-api/api/token/", data=json.dumps(account_data),headers=headers)
	token=response_obj.json()['data']['access']
	print(token)
	if response_obj.status_code >= 200 and response_obj.status_code<300:
		try:
			
				log=""
				list_id_pdf_file = request.data["list_id_pdf_file"]
				
				list_id=list_id_pdf_file.split(",")
				accountCate=[]
				
				for i in list_id:
					account = str(pdfFile.objects.get(pk=int(i)).account)
					if account not in accountCate:
						accountCate.append(account)
				for account in accountCate:
					listfile=[]
					for i in list_id:
						if account == str(pdfFile.objects.get(pk=int(i)).account):
							pdf=pdfFile.objects.get(pk=int(i))	
							if pdf.sended == False:
								pdffile=pdfFile.objects.get(pk=int(i)).slaveFile
								linkfile=settings.URL+"/"+str(pdffile)
								tple=detect_position(str(pdffile))
								print(tple)
								data_send ={
									"pdf_url":linkfile,
									"sign_pos": "{}x{}".format(round(tple[0]),round(tple[1])),
									"contact": "thach.nguyenphamngoc@kcc.com",
									"reason": "sign contract",
									"page_number":tple[2]
								}							
								headers2 = { 'Content-Type':'application/json', 
								'Authorization': 'Bearer ' + token }
								print(data_send)
								response_obj2 = requests.post(r"https://api-testing.pvs.com.vn/e-invoice-api/api/ca-sign/sign-pdf/84", data=json.dumps(data_send), headers=headers2)				
								if response_obj2.status_code  >= 200 and response_obj2.status_code<300:
									log+="{}:success ".format(str(pdffile).replace("documents/slavefiles/",""))
									binarytext=response_obj2.content
									with open("file.pdf","wb") as file:
										file.write(binarytext)
									with open ("file.pdf","rb") as file:
										name="ThưThôngBáo_{}_{}.pdf".format(account,str(datetime.now().date()))
										newpdf=pdf.slaveFile.save(name,File(file))
										pdf.sended=True
										pdf.signed=True
										pdf.sendingTime=datetime.now()
										pdf.save()
										fileurl= settings.URL+"/"+str(pdf.slaveFile)
									listfile.append(fileurl)
								if response_obj2.status_code  >= 300 and response_obj2.status_code <= 500:
									try:
										log+=str(response_obj2.json()['message'])
									except:
										log+=str(response_obj2.json()['message'])
										continue
								
							else:
								pdffile=pdfFile.objects.get(pk=int(i)).slaveFile
								linkfile=settings.URL+"/"+str(pdffile)
								listfile.append(linkfile)
								log+="{}:success ".format(str(pdffile).replace("documents/slavefiles/",""))
							if listfile != []:
								for email in pdf.emailExtracted.all():
									print(listfile)
									print(email)
									send_email.send_noti_to_partner_sign_by_email(listfile,str(email))
							else:
								log+=("listfile=[]")
				return Response({"log":log}, status=status.HTTP_200_OK)
		except:
			err_mess = sys.exc_info()[0].__name__ + ": "+ str(sys.exc_info()[1])
			print(err_mess)
			return Response({"err":err_mess},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	else :
		return Response(response_obj.json()['message'], status=status.HTTP_500_INTERNAL_SERVER_ERROR)
@api_view(["POST"])
def deleteFile(request):
	print(request.data)
	if request.data["list_id_pdf_file"]:
		list_id_pdf_file = request.data["list_id_pdf_file"]
		list_id=list_id_pdf_file.split(",")
		print(list_id)
		for id in list_id:
			fileWillBeDel=pdfFile.objects.get(pk=int(id))
			print(fileWillBeDel)
			fileWillBeDel.delete()
	return Response({"msg":"delete success"}, status=status.HTTP_200_OK)

@api_view(["POST"])
def deleteExcelFile(request):
	print(request.data)
	if request.data["list_id_excel_file"]:
		list_id_pdf_file = request.data["list_id_excel_file"]
		list_id=list_id_pdf_file.split(",")
		print(list_id)
		for id in list_id:
			fileWillBeDel=document.objects.get(pk=int(id))
			print(fileWillBeDel)
			fileWillBeDel.delete()
	return Response({"msg":"delete success"}, status=status.HTTP_200_OK)


def export_hnk_ticket_excel(from_date, to_date):
	col_names = ["","file","creator","createtime","sendingtime"]
	actlogs = pdfFile.objects.filter(createdTime__date__gte=from_date,createdTime__date__lte=to_date)|pdfFile.objects.filter(sendingTime__date__gte=from_date,createdTime__date__lte=to_date)

	wb = openpyxl.Workbook()
	wb.iso_dates = True
	ws = wb.create_sheet(title="actlog")

	# create title
	for col in range(1, len(col_names)):
		_ = ws.cell(column=col,row=1,value=col_names[col])
	
	alignment=Alignment(horizontal='general')
	normal_format = NamedStyle(name="normal",alignment=alignment)
	datetime_format = NamedStyle(name="datetime",number_format="DD/MMM/YYYY h:mm",alignment=alignment)
	date_format = NamedStyle(name="date",number_format="DD/MMM/YYYY",alignment=alignment)
	# create content
	# now = pytz.utc.localize(datetime.utcnow())
	# now = now.replace(tzinfo=None)
	# print(now)
	for row, ticket in enumerate(actlogs, start=2):
		for col in range(1,5):		
			if col_names[col] == "file" :
				if ticket.slaveFile:
					c = ws.cell(column=col,row=row,value=str(ticket.slaveFile).replace("documents/slavefiles/","") )
				else:
					c = ws.cell(column=col,row=row,value="")
				c.style = datetime_format
				
			elif col_names[col] == "creator":
				c = ws.cell(column=col,row=row,value=str(ticket.creator))
				c.style = date_format
			elif col_names[col] == "createtime":
				if ticket.createdTime:
					c = ws.cell(column=col,row=row,value=str(ticket.createdTime.date()))
				else:
					c = ws.cell(column=col,row=row,value="")
				c.style = normal_format
			elif col_names[col] == "sendingtime":
				if ticket.sendingTime:
					c = ws.cell(column=col,row=row,value=str(ticket.sendingTime.date()))
					c.style = normal_format
				else:
					c = ws.cell(column=col,row=row,value="")
	

	response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment;filename={}_{}.xlsx'.format(from_date,to_date)

	return response

@api_view(["POST"])
def update_profile(request):
	user = request.user
	profile = Profile.objects.get(user=user)
	profile.full_name = request.data['full_name']
	profile.phone_number = request.data['phone_number']
	profile.email = request.data['email']
	profile.company_name = request.data['company_name']
	profile.position= request.data['position']
	profile.address = request.data['address']
	profile.save()
	print(request.data)
	return Response({"msg":"success"}, status=status.HTTP_200_OK)